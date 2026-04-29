from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def _flatten_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if isinstance(out.columns, pd.MultiIndex):
        new_cols = []
        for col in out.columns:
            parts = [str(x) for x in col if x not in (None, "", "nan")]
            new_cols.append(" | ".join(parts) if parts else "")
        out.columns = new_cols
    return out


def write_xlsx(df: pd.DataFrame, output_file: Path, metadata: dict[str, Any]) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    flat = _flatten_columns(df)
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        flat.to_excel(writer, sheet_name="export", index=False)
        meta_rows = [[k, _format_meta_value(v)] for k, v in metadata.items()]
        pd.DataFrame(meta_rows, columns=["key", "value"]).to_excel(writer, sheet_name="metadata", index=False)
    _beautify(output_file)


def write_csv(df: pd.DataFrame, output_file: Path) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    _flatten_columns(df).to_csv(output_file, index=False)


def write_metadata(path: Path, metadata: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [f"{k}: {_format_meta_value(v)}" for k, v in metadata.items()]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _format_meta_value(v: Any) -> str:
    if isinstance(v, (list, tuple)):
        return ", ".join(map(str, v))
    return "" if v is None else str(v)


def _beautify(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="EAF2F8")
    meta_fill = PatternFill("solid", fgColor="F4F6F7")
    thin = Side(style="thin", color="D9E2EC")
    border = Border(bottom=thin)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        max_row = ws.max_row
        max_col = ws.max_column
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = header_fill if ws.title == "export" else meta_fill
            cell.border = border
        for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=False)
        for col_idx in range(1, max_col + 1):
            letter = get_column_letter(col_idx)
            values = [ws.cell(r, col_idx).value for r in range(1, min(max_row, 80) + 1)]
            max_len = max([len(str(v)) if v is not None else 0 for v in values] + [8])
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 34)
        ws.row_dimensions[1].height = 34 if ws.title == "export" else 22
    wb.save(path)
