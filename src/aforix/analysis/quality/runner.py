from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

from aforix.analysis.quality.config import load_quality_config
from aforix.analysis.quality.metrics import compute_cg_from_weights, find_tq_column

FILE_RE = re.compile(
    r"^(?P<point>(?:P)?\d+)_Points_(?P<date>\d{8})(?:_(?P<time>\d{6}))?\.csv$",
    re.IGNORECASE,
)


def run_quality_metrics(
    config_path: str | Path,
    aggregation: str = "daily",
    points: list[str] | None = None,
    months: list[str] | None = None,
    all_months: bool = False,
) -> Path:
    qc = load_quality_config(config_path)

    if not qc.enabled:
        raise RuntimeError("quality_metrics disabled in config")

    if aggregation not in {"measurement", "daily", "monthly"}:
        raise ValueError("aggregation must be one of: measurement, daily, monthly")

    wanted_points = {_normalize_point(p) for p in points or []}
    wanted_months = {_normalize_month(m) for m in months or []}

    norm_dir = qc.nivus.normalized_points
    raw_dir = qc.nivus.raw_points

    if not norm_dir.exists():
        raise FileNotFoundError(f"Normalized Points dir not found: {norm_dir}")
    if not raw_dir.exists():
        raise FileNotFoundError(f"Raw Points dir not found: {raw_dir}")

    records: list[dict[str, object]] = []
    logs: list[dict[str, object]] = []

    for norm_file in sorted(norm_dir.glob("*.csv")):
        meta = _parse_measurement_filename(norm_file.name)
        if meta is None:
            logs.append({"file": norm_file.name, "status": "skipped", "reason": "filename_not_recognized"})
            continue

        point_id = _normalize_point(meta["point"])
        point_label = _format_point_label(meta["point"])
        month_id = meta["date"][:6]

        if wanted_points and point_id not in wanted_points:
            continue
        if not all_months and wanted_months and month_id not in wanted_months:
            continue

        raw_file = raw_dir / norm_file.name
        if not raw_file.exists():
            logs.append({"file": norm_file.name, "status": "missing_raw", "reason": f"Raw file not found: {raw_file}"})
            continue

        try:
            df_norm = pd.read_csv(norm_file)
            df_raw = pd.read_csv(raw_file)

            weight_col = qc.nivus.weight_column
            if weight_col not in df_norm.columns:
                raise ValueError(f"Missing weight column '{weight_col}' in normalized Points")

            tq_col = find_tq_column(df_raw)

            if len(df_norm) != len(df_raw):
                raise ValueError(
                    f"Normalized and raw Points length mismatch: normalized={len(df_norm)}, raw={len(df_raw)}"
                )

            cg = compute_cg_from_weights(df_norm[weight_col], df_raw[tq_col])
            period = _period_from_meta(meta, aggregation)

            records.append(
                {
                    "point": point_label,
                    "date": meta["date"],
                    "time": meta["time"],
                    "period": period,
                    "cg": cg,
                    "file": norm_file.name,
                }
            )
            logs.append(
                {
                    "file": norm_file.name,
                    "status": "ok",
                    "point": point_label,
                    "date": meta["date"],
                    "time": meta["time"],
                    "tq_column": tq_col,
                    "weight_column": weight_col,
                }
            )

        except Exception as exc:
            logs.append(
                {
                    "file": norm_file.name,
                    "status": "error",
                    "point": point_label,
                    "date": meta["date"],
                    "time": meta["time"],
                    "error": str(exc),
                }
            )

    out_dir = qc.paths.output_root / datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir.mkdir(parents=True, exist_ok=True)

    results_df = pd.DataFrame(records)
    log_df = pd.DataFrame(logs)

    if records:
        cg_table = results_df.pivot_table(
            index="point",
            columns="period",
            values="cg",
            aggfunc="mean",
        )
        cg_table = cg_table.reindex(sorted(cg_table.columns), axis=1)
        cg_table = cg_table.sort_index(key=lambda idx: idx.map(_sort_key))
    else:
        cg_table = pd.DataFrame()

    results_df.to_csv(out_dir / "cg_measurements.csv", index=False)
    log_df.to_csv(out_dir / "cg_log.csv", index=False)

    excel_path = out_dir / "CG.xlsx"
    _write_excel_report(excel_path, cg_table, log_df, results_df, aggregation)

    return out_dir


def discover_available_filters(config_path: str | Path) -> tuple[list[str], list[str]]:
    qc = load_quality_config(config_path)
    norm_dir = qc.nivus.normalized_points
    points: set[str] = set()
    months: set[str] = set()

    if not norm_dir.exists():
        return [], []

    for norm_file in sorted(norm_dir.glob("*.csv")):
        meta = _parse_measurement_filename(norm_file.name)
        if meta is None:
            continue
        points.add(_format_point_label(meta["point"]))
        months.add(meta["date"][:6])

    return sorted(points, key=_sort_key), sorted(months)


def _write_excel_report(
    excel_path: Path,
    cg_table: pd.DataFrame,
    log_df: pd.DataFrame,
    results_df: pd.DataFrame,
    aggregation: str,
) -> None:
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        cg_table.to_excel(writer, sheet_name="CG")
        log_df.to_excel(writer, sheet_name="Log", index=False)
        results_df.to_excel(writer, sheet_name="Measurements", index=False)

        wb = writer.book
        ws = wb["CG"]
        chart_ws = wb.create_sheet("Charts")

        _autosize(ws)
        _autosize(wb["Log"])
        _autosize(wb["Measurements"])

        if not cg_table.empty:
            _add_line_charts(ws, chart_ws, aggregation)


def _add_line_charts(cg_ws, chart_ws, aggregation: str) -> None:
    max_row = cg_ws.max_row
    max_col = cg_ws.max_column

    if max_row < 2 or max_col < 2:
        return

    categories = Reference(cg_ws, min_col=2, max_col=max_col, min_row=1)

    chart_row_anchor = 1
    for row in range(2, max_row + 1):
        point_label = cg_ws.cell(row=row, column=1).value
        chart = LineChart()
        chart.title = f"CG(%) - {point_label}"
        chart.y_axis.title = "CG(%)"
        chart.x_axis.title = "Fecha" if aggregation != "monthly" else "Mes"
        chart.height = 7
        chart.width = 18

        data = Reference(cg_ws, min_col=2, max_col=max_col, min_row=row, max_row=row)
        chart.add_data(data, from_rows=True, titles_from_data=False)
        chart.set_categories(categories)

        chart_ws.add_chart(chart, f"A{chart_row_anchor}")
        chart_row_anchor += 15


def _autosize(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)


def _parse_measurement_filename(name: str) -> dict[str, str] | None:
    match = FILE_RE.match(name)
    if not match:
        return None
    return {
        "point": match.group("point"),
        "date": match.group("date"),
        "time": match.group("time") or "000000",
    }


def _period_from_meta(meta: dict[str, str], aggregation: str) -> str:
    if aggregation == "monthly":
        return meta["date"][:6]
    if aggregation == "measurement":
        return f"{meta['date']}_{meta['time']}"
    return meta["date"]


def _normalize_point(value: str) -> str:
    return str(value).strip().upper().replace("P", "")


def _format_point_label(value: str) -> str:
    value = str(value).strip().upper()
    return value if value.startswith("P") else value


def _sort_key(value: object) -> int:
    return int(_normalize_point(str(value)))


def _normalize_month(value: str) -> str:
    value = str(value).strip()
    if not re.fullmatch(r"\d{6}", value):
        raise ValueError(f"Invalid month '{value}'. Expected YYYYMM.")
    return value
