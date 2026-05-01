from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, ScatterChart, Series
from openpyxl.utils.dataframe import dataframe_to_rows


def safe_save_workbook(wb: Workbook, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(path)
        return path
    except PermissionError:
        for i in range(2, 100):
            alt = path.with_name(f"{path.stem}_v{i}{path.suffix}")
            try:
                wb.save(alt)
                return alt
            except PermissionError:
                continue
        raise


def add_pair_sheet(
    wb: Workbook,
    sheet_name: str,
    data: pd.DataFrame,
    metrics: dict[str, Any],
    *,
    x_col: str,
    y_col: str,
    pred_col: str,
    time_col: str,
    x_label: str,
    y_label: str,
) -> None:
    ws = wb.create_sheet(sheet_name[:31])

    for row in dataframe_to_rows(data, index=False, header=True):
        ws.append(row)

    row_off = len(data) + 3
    for idx, (key, value) in enumerate(metrics.items()):
        ws.cell(row=row_off + idx, column=1, value=key)
        ws.cell(row=row_off + idx, column=2, value=value)

    headers = list(data.columns)
    nrows = len(data) + 1
    x_idx = headers.index(x_col) + 1
    y_idx = headers.index(y_col) + 1
    pred_idx = headers.index(pred_col) + 1
    time_idx = headers.index(time_col) + 1

    scatter = ScatterChart()
    scatter.title = f"{y_label} vs {x_label}"
    scatter.x_axis.title = x_label
    scatter.y_axis.title = y_label
    scatter.x_axis.majorGridlines = None
    scatter.y_axis.majorGridlines = None

    x_vals = Reference(ws, min_col=x_idx, min_row=2, max_row=nrows)
    y_vals = Reference(ws, min_col=y_idx, min_row=2, max_row=nrows)
    pts = Series(y_vals, x_vals, title="Observed")
    pts.marker.symbol = "circle"
    pts.graphicalProperties.line.noFill = True
    scatter.series.append(pts)

    aux_col = len(headers) + 2
    ws.cell(row=1, column=aux_col, value="y=x")
    for r in range(2, nrows + 1):
        ws.cell(row=r, column=aux_col, value=ws.cell(row=r, column=x_idx).value)
    one_to_one = Reference(ws, min_col=aux_col, min_row=2, max_row=nrows)
    s_11 = Series(one_to_one, x_vals, title="1:1")
    s_11.marker = None
    scatter.series.append(s_11)

    pred_vals = Reference(ws, min_col=pred_idx, min_row=2, max_row=nrows)
    reg = Series(pred_vals, x_vals, title="Regression")
    reg.marker = None
    scatter.series.append(reg)
    ws.add_chart(scatter, "I2")

    line = LineChart()
    line.title = "Time Series"
    line.x_axis.title = "Time"
    line.y_axis.title = "Flow [l/s]"
    cats = Reference(ws, min_col=time_idx, min_row=2, max_row=nrows)
    line.set_categories(cats)
    line.series.append(Series(Reference(ws, min_col=x_idx, min_row=2, max_row=nrows), title=x_label))
    line.series.append(Series(Reference(ws, min_col=y_idx, min_row=2, max_row=nrows), title=y_label))
    ws.add_chart(line, "I20")


def write_summary_sheet(wb: Workbook, name: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(name[:31])
    if not rows:
        ws.append(["No data generated"])
        return
    df = pd.DataFrame(rows)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
