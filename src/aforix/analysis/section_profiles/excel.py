from __future__ import annotations

from pathlib import Path
from typing import Dict, Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, BarChart, Reference, Series
from openpyxl.styles import Font


def write_excel(
    output_path: Path,
    sheets: list[dict],
    *,
    x_axis: str,
    y_axis: str,
    chart_type: str,
    excel_cfg: Dict[str, Any] | None = None,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    excel_cfg = excel_cfg or {}

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        if excel_cfg.get('include_readme', True):
            _readme_table(output_path, x_axis=x_axis, y_axis=y_axis, chart_type=chart_type).to_excel(
                writer, sheet_name='README', index=False
            )
        if excel_cfg.get('include_index', True):
            _index_table(sheets).to_excel(writer, sheet_name='Index', index=False)

        for sh in sheets:
            name = sh['sheet_name']
            df = sh['data']
            summary = sh['summary']

            sum_df = pd.DataFrame(list(summary.items()), columns=['item', 'value'])
            sum_df.to_excel(writer, sheet_name=name, index=False, startrow=0)

            start_row = len(sum_df) + 2
            df.to_excel(writer, sheet_name=name, index=False, startrow=start_row)

    wb = load_workbook(output_path)

    for ws in wb.worksheets:
        _format_sheet(ws)

    for sh in sheets:
        ws = wb[sh['sheet_name']]
        df = sh['data']

        if df.empty:
            continue

        ws._charts = []

        start_row = len(sh['summary']) + 3
        end_row = start_row + len(df) - 1

        headers = list(df.columns)
        if x_axis not in headers or y_axis not in headers:
            continue

        x_col = headers.index(x_axis) + 1
        y_col = headers.index(y_axis) + 1

        x_ref = Reference(ws, min_col=x_col, min_row=start_row + 1, max_row=end_row + 1)
        y_ref = Reference(ws, min_col=y_col, min_row=start_row + 1, max_row=end_row + 1)

        if chart_type == 'bar':
            chart = BarChart()
            chart.add_data(y_ref, titles_from_data=False)
            chart.set_categories(x_ref)
        else:
            chart = ScatterChart()
            series = Series(y_ref, x_ref, title=y_axis)
            chart.series.append(series)

        chart.title = f"{y_axis} vs {x_axis}"
        chart.x_axis.title = x_axis
        chart.y_axis.title = y_axis

        anchor = excel_cfg.get('chart_anchor', 'H2')
        ws.add_chart(chart, anchor)

    wb.save(output_path)
    return output_path


def _readme_table(output_path: Path, *, x_axis: str, y_axis: str, chart_type: str) -> pd.DataFrame:
    rows = [
        {'item': 'report', 'value': 'Section profiles analysis'},
        {'item': 'output_file', 'value': str(output_path)},
        {'item': 'source', 'value': 'database/normalized/{instrument}/Points/*.csv'},
        {'item': 'x_axis', 'value': x_axis},
        {'item': 'y_axis', 'value': y_axis},
        {'item': 'chart_type', 'value': chart_type},
        {'item': 'structure', 'value': 'One worksheet per measurement plus README and Index sheets.'},
        {'item': 'charts', 'value': 'Native Excel charts generated with openpyxl.'},
    ]
    return pd.DataFrame(rows)


def _index_table(sheets: list[dict]) -> pd.DataFrame:
    rows = []
    for sh in sheets:
        summary = sh.get('summary', {})
        rows.append(
            {
                'sheet_name': sh.get('sheet_name'),
                'station_id': summary.get('station_id'),
                'measurement_date': summary.get('measurement_date'),
                'measurement_time': summary.get('measurement_time'),
                'instrument': summary.get('instrument'),
                'instrument_code': summary.get('instrument_code'),
                'n_rows': summary.get('n_rows'),
                'source_file': summary.get('source_file'),
            }
        )
    return pd.DataFrame(rows)


def _format_sheet(ws) -> None:
    ws.freeze_panes = 'A2'
    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = Font(bold=True)
    for column_cells in ws.columns:
        max_len = 0
        col_letter = column_cells[0].column_letter
        for cell in column_cells[:250]:
            value = '' if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 55)
